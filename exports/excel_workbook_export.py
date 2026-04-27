from __future__ import annotations

from typing import Dict, List, Set

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from core.project import ProjectData
from core.timeline import derive_actor_timelines
from core.final_output import compute_final_mic_numbering, compute_wire_numbering


# -----------------------------
# Helpers
# -----------------------------

def _abbreviate_name(full_name: str) -> str:
    parts = (full_name or "").strip().split()
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0]
    return f"{parts[0]} {parts[-1][0]}"


def _autosize_columns(ws, min_width=10, max_width=45, factor=1.25):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        width = int(max_len * factor) + 2
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, width))


def _ordered_scenes(project: ProjectData):
    return sorted(
        project.scenes,
        key=lambda s: (
            getattr(s, "start_page", 0) or 0,
            getattr(s, "end_page", 0) or 0,
            getattr(s, "id", 0) or 0,
        ),
    )


def _scene_header(scene) -> str:
    name = (scene.name or "").strip()
    sp = int(getattr(scene, "start_page", 0) or 0)
    ep = int(getattr(scene, "end_page", 0) or 0)
    if sp and ep and sp != ep:
        return f"{name} Pg {sp}-{ep}"
    if sp:
        return f"{name} Pg {sp}"
    return name


# -----------------------------
# Base borders / alignments
# -----------------------------

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


# -----------------------------
# Colors (ARGB)
# -----------------------------

FILL_GREEN = PatternFill("solid", fgColor="FF92D050")
FILL_RED = PatternFill("solid", fgColor="FFFF0000")
FILL_YELLOW = PatternFill("solid", fgColor="FFFFFF00")

FILL_WHITE = PatternFill("solid", fgColor="FFFFFFFF")
FILL_NONE = PatternFill()


# -----------------------------
# Mic List styling
# -----------------------------

MIC_LIST_HEADER_FONT = Font(name="Aptos Narrow", bold=True, size=20)
MIC_LIST_BODY_FONT = Font(name="Aptos Narrow", bold=True, size=14)

MIC_LIST_HEADER_FILL = FILL_GREEN
MIC_LIST_BODY_FILL = FILL_YELLOW


# -----------------------------
# Sharing styling
# -----------------------------

SHARING_HEADER_FONT = Font(name="Aptos Narrow", bold=True, size=20)
SHARING_BODY_FONT = Font(name="Aptos Narrow", bold=True, size=14)

SHARING_HEADER_FILL = FILL_GREEN
SHARING_BODY_FILL = FILL_YELLOW


# -----------------------------
# Mic Plot styling
# -----------------------------

MIC_PLOT_FONT = Font(name="Aptos Narrow", bold=True, size=11)
MIC_PLOT_HEADER_FILL = FILL_YELLOW
MIC_PLOT_PACK_FILL = FILL_YELLOW
MIC_PLOT_HEADER_HEIGHT = 105.0


# -----------------------------
# Scenes styling (your final spec)
# -----------------------------

SCENES_FONT = Font(name="Aptos Narrow", bold=True, size=11)
SCENES_HEADER_FILL = FILL_YELLOW
SCENES_INFO_FILL = FILL_YELLOW   # Wire / Cast / Pack columns
SCENES_SCENE_FILL = FILL_WHITE   # ALL scene cells, X or empty


# -----------------------------
# Override-aware context builder
# -----------------------------

def _compute_override_context(project: ProjectData, assignments: list):
    """
    Returns:
      scenes (page-ordered),
      per_scene_by_mic (mic -> scene_idx -> set(actors)),
      actor_indices (actor -> list of scene indices they appear in),
      final_mics,
      wire_res
    """
    scenes = _ordered_scenes(project)

    original_scenes = project.scenes
    project.scenes = scenes
    try:
        overrides = getattr(project, "mic_scene_overrides", {}) or {}

        # actor -> timeline indices
        timelines = derive_actor_timelines(project, include_uncast=True)
        actor_indices: Dict[str, List[int]] = {}
        actor_index_sets: Dict[str, Set[int]] = {}
        for actor, tl in timelines.items():
            idxs = [i for i in tl.indices if 0 <= i < len(scenes)]
            actor_indices[actor] = idxs
            actor_index_sets[actor] = set(idxs)

        # numbering + wire numbering (these remain plan-based; overrides don't change wire assignment)
        final_mics = compute_final_mic_numbering(
            project=project,
            assignments=assignments,
            grouping_mode=getattr(project, "grouping_mode", "none"),
            actor_groups=getattr(project, "actor_groups", None),
            character_groups=getattr(project, "character_groups", None),
        )

        wire_res = compute_wire_numbering(
            project=project,
            assignments=assignments,
            final_mic_numbering=final_mics,
            grouping_mode=getattr(project, "grouping_mode", "none"),
            actor_groups=getattr(project, "actor_groups", None),
            character_groups=getattr(project, "character_groups", None),
        )

        # forced map: scene_idx -> mic_num -> set(actors)
        forced: Dict[int, Dict[int, Set[str]]] = {}
        for (actor, scene_i), chosen_mic in overrides.items():
            if actor not in actor_index_sets:
                continue
            if scene_i not in actor_index_sets[actor]:
                continue
            forced.setdefault(int(scene_i), {}).setdefault(int(chosen_mic), set()).add(actor)

        # per_scene_by_mic: mic -> scene -> set(actors)
        per_scene_by_mic: Dict[int, Dict[int, Set[str]]] = {}

        mic_nums = {int(m.mic_number) for m in assignments}
        for mic in assignments:
            mic_num = int(mic.mic_number)
            per_scene: Dict[int, Set[str]] = {}

            # base: assigned actors unless overridden away
            for actor in (mic.actors or []):
                for i in actor_indices.get(actor, []):
                    chosen = overrides.get((actor, i), None)
                    if chosen is not None and int(chosen) != mic_num:
                        continue
                    per_scene.setdefault(i, set()).add(actor)

            # forced-in: actors overridden to this mic
            for i, mp in forced.items():
                addset = mp.get(mic_num)
                if addset:
                    per_scene.setdefault(i, set()).update(addset)

            per_scene_by_mic[mic_num] = per_scene

        # ignore overrides to non-existent mic rows (not in current plan)
        # Timeline UI allows any mic; export only includes mics that exist in assignments.

        return scenes, per_scene_by_mic, actor_indices, final_mics, wire_res
    finally:
        project.scenes = original_scenes


# -----------------------------
# Mic Plot helpers (ignore blanks)
# -----------------------------

def _prev_non_blank(active: List[str], idx: int) -> str:
    i = idx - 1
    while i >= 0:
        if active[i]:
            return active[i]
        i -= 1
    return ""


def _next_non_blank(active: List[str], idx: int) -> str:
    i = idx + 1
    while i < len(active):
        if active[i]:
            return active[i]
        i += 1
    return ""


# -----------------------------
# Sheet builders (override-aware)
# -----------------------------

def _build_mic_plot(wb: Workbook, project: ProjectData, assignments: list):
    ws = wb.create_sheet("Mic Plot")
    scenes, per_scene_by_mic, _actor_indices, final_mics, _wire_res = _compute_override_context(project, assignments)

    # Pack order must match Final Output Preview (final mic numbering)
    mics_sorted = sorted(assignments, key=lambda a: int(final_mics.get(int(a.mic_number), int(a.mic_number))))

    headers = ["Mic Pack"] + [_scene_header(s) for s in scenes]
    ws.append(headers)

    ws.row_dimensions[1].height = MIC_PLOT_HEADER_HEIGHT
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = MIC_PLOT_FONT
        cell.alignment = CENTER_WRAP
        cell.fill = MIC_PLOT_HEADER_FILL
        cell.border = BORDER

    scene_count = len(scenes)

    for r, mic in enumerate(mics_sorted, start=2):
        mic_num = int(mic.mic_number)
        final_pack = int(final_mics.get(mic_num, mic_num))

        cpack = ws.cell(row=r, column=1, value=final_pack)
        cpack.font = MIC_PLOT_FONT
        cpack.alignment = CENTER
        cpack.fill = MIC_PLOT_PACK_FILL
        cpack.border = BORDER

        per_scene = per_scene_by_mic.get(mic_num, {})

        # Build active actor per scene (choose one if multiple)
        active: List[str] = []
        for si in range(scene_count):
            aset = per_scene.get(si, set())
            if not aset:
                active.append("")
            else:
                active.append(sorted(aset, key=str.lower)[0])

        # Render cells with swap coloring
        for c_idx, actor in enumerate(active, start=2):
            idx = c_idx - 2
            cell = ws.cell(row=r, column=c_idx, value=_abbreviate_name(actor) if actor else "")
            cell.font = MIC_PLOT_FONT
            cell.alignment = CENTER
            cell.border = BORDER
            cell.fill = FILL_NONE

            if not actor:
                continue

            prev_actor = _prev_non_blank(active, idx)
            next_actor = _next_non_blank(active, idx)

            start = (prev_actor != actor)
            end = (next_actor != actor)

            if start and end:
                cell.fill = FILL_YELLOW
            elif start:
                cell.fill = FILL_GREEN
            elif end:
                cell.fill = FILL_RED

    _autosize_columns(ws, min_width=8, max_width=32)


def _build_sharing(wb: Workbook, project: ProjectData, assignments: list):
    ws = wb.create_sheet("Sharing")
    scenes, per_scene_by_mic, _actor_indices, final_mics, _wire_res = _compute_override_context(project, assignments)

    mics_sorted = sorted(assignments, key=lambda a: int(final_mics.get(int(a.mic_number), int(a.mic_number))))

    # For sharing: actors who EVER use this mic (including via overrides)
    mic_to_actorlist: Dict[int, List[str]] = {}
    max_sharers = 0
    for mic in mics_sorted:
        mic_num = int(mic.mic_number)
        aset_all: Set[str] = set()
        for _si, aset in per_scene_by_mic.get(mic_num, {}).items():
            aset_all.update(aset)
        alist = sorted(aset_all, key=str.lower)
        mic_to_actorlist[mic_num] = alist
        max_sharers = max(max_sharers, len(alist))

    headers = ["Pack #"] + [f"Cast Member {i}" for i in range(1, max_sharers + 1)]
    ws.append(headers)

    ws.row_dimensions[1].height = 26.25
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = SHARING_HEADER_FONT
        cell.alignment = CENTER if c == 1 else LEFT
        cell.fill = SHARING_HEADER_FILL
        cell.border = BORDER

    for r, mic in enumerate(mics_sorted, start=2):
        ws.row_dimensions[r].height = 18.75
        mic_num = int(mic.mic_number)
        final_pack = int(final_mics.get(mic_num, mic_num))

        c1 = ws.cell(row=r, column=1, value=final_pack)
        c1.font = SHARING_BODY_FONT
        c1.alignment = CENTER
        c1.fill = SHARING_BODY_FILL
        c1.border = BORDER

        alist = mic_to_actorlist.get(mic_num, [])
        for i in range(max_sharers):
            val = alist[i] if i < len(alist) else ""
            cell = ws.cell(row=r, column=2 + i, value=val)
            cell.font = SHARING_BODY_FONT
            cell.alignment = LEFT
            cell.fill = SHARING_BODY_FILL
            cell.border = BORDER

    _autosize_columns(ws, min_width=10, max_width=45)


def _build_mic_list(wb: Workbook, project: ProjectData, assignments: list):
    ws = wb.create_sheet("Mic List")
    scenes, per_scene_by_mic, _actor_indices, final_mics, wire_res = _compute_override_context(project, assignments)

    ws.append(["Wire #", "Pack #", "Cast Member"])

    for c, align in [(1, CENTER_WRAP), (2, CENTER_WRAP), (3, LEFT)]:
        cell = ws.cell(row=1, column=c)
        cell.font = MIC_LIST_HEADER_FONT
        cell.alignment = align
        cell.fill = MIC_LIST_HEADER_FILL
        cell.border = BORDER

    # actor -> set(internal mics ever used)
    actor_to_mics: Dict[str, Set[int]] = {}
    for mic_num, per_scene in per_scene_by_mic.items():
        for _si, aset in per_scene.items():
            for actor in aset:
                actor_to_mics.setdefault(actor, set()).add(int(mic_num))

    # Build rows: (wire, final_pack, actor)
    rows = []
    for actor, mset in actor_to_mics.items():
        wire = wire_res.actor_to_wire.get(actor)
        if wire is None:
            continue
        for internal_mic in sorted(mset):
            final_pack = int(final_mics.get(internal_mic, internal_mic))
            rows.append((wire, final_pack, actor))

    rows.sort(key=lambda r: (r[0], r[1], r[2].lower()))

    for r, (wire, pack, actor) in enumerate(rows, start=2):
        for c, val, align in [
            (1, wire, CENTER),
            (2, pack, CENTER),
            (3, actor, LEFT),
        ]:
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = MIC_LIST_BODY_FONT
            cell.alignment = align
            cell.fill = MIC_LIST_BODY_FILL
            cell.border = BORDER

    _autosize_columns(ws, min_width=10, max_width=45)


def _build_scenes(wb: Workbook, project: ProjectData, assignments: list):
    ws = wb.create_sheet("Scenes")
    scenes, per_scene_by_mic, actor_indices, final_mics, wire_res = _compute_override_context(project, assignments)

    headers = ["Wire #", "Cast Member", "Mic Pack"] + [_scene_header(s) for s in scenes]
    ws.append(headers)

    ws.row_dimensions[1].height = MIC_PLOT_HEADER_HEIGHT
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = SCENES_FONT
        cell.alignment = CENTER_WRAP
        cell.fill = SCENES_HEADER_FILL
        cell.border = BORDER

    # actor -> set(internal mics ever used)
    actor_to_mics: Dict[str, Set[int]] = {}
    for mic_num, per_scene in per_scene_by_mic.items():
        for _si, aset in per_scene.items():
            for actor in aset:
                actor_to_mics.setdefault(actor, set()).add(int(mic_num))

    # Build expanded rows: one row per (wire, actor, final_pack)
    expanded_rows: list[tuple[int, str, int]] = []
    for actor, wire in sorted(wire_res.actor_to_wire.items(), key=lambda kv: kv[1]):
        internal_mics = sorted(actor_to_mics.get(actor, set()))
        final_packs = sorted({int(final_mics.get(m, m)) for m in internal_mics})

        # If actor has no mic usage (should be rare), still output one row with blank pack
        if not final_packs:
            expanded_rows.append((int(wire), actor, 0))
        else:
            for pack in final_packs:
                expanded_rows.append((int(wire), actor, int(pack)))

    # Sort: Wire # first, then Mic Pack for duplicates, then Actor name
    expanded_rows.sort(key=lambda r: (r[0], r[2], r[1].lower()))

    # Render each expanded row
    for r, (wire, actor, pack) in enumerate(expanded_rows, start=2):
        pack_text = "" if pack == 0 else str(pack)

        # First three columns (yellow)
        for c, val, align in [
            (1, wire, CENTER),
            (2, _abbreviate_name(actor), LEFT),
            (3, pack_text, CENTER),
        ]:
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = SCENES_FONT
            cell.alignment = align
            cell.fill = SCENES_INFO_FILL
            cell.border = BORDER

        # Scene participation: X if actor appears in that scene (same for each duplicated row)
        aset_idx = set(actor_indices.get(actor, []))
        for c, si in enumerate(range(len(scenes)), start=4):
            cell = ws.cell(row=r, column=c, value=("X" if si in aset_idx else ""))
            cell.font = SCENES_FONT
            cell.alignment = CENTER
            cell.fill = SCENES_SCENE_FILL   # ALL white (X or blank)
            cell.border = BORDER

    _autosize_columns(ws, min_width=8, max_width=32)


# -----------------------------
# Public entry point
# -----------------------------

def export_show_workbook(path: str, project: ProjectData, assignments: list):
    wb = Workbook()
    wb.remove(wb.active)

    _build_mic_plot(wb, project, assignments)
    _build_sharing(wb, project, assignments)
    _build_mic_list(wb, project, assignments)
    _build_scenes(wb, project, assignments)

    wb.save(path)